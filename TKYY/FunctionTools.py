import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font,PatternFill,Alignment,Side,Border
import paramiko
import datetime
import time
import logging
import re
import requests
import os
import zipfile
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from email.utils import formataddr
import smtplib



# Jul.11.2022 : Add a different background color according to the cell value, such as red, orange for warning and string function.



def percent_to_int(string):
    if '%' in string:
        try:
            ints = int(string.strip('%').split('.')[0])
            return ints
        except:
            return

    if 'w' in string:
        try:
            ints = int(string.split('w')[0])
            return ints
        except:
            return



def change_xl_style(sheet_name,file_name):

    #wb = openpyxl.Workbook()
    # that will rewrite exist excel file
    wb = openpyxl.load_workbook(file_name)

    ws = wb.active
    ws.title = sheet_name


    font1 = Font(
        name="Calibri",
        size=11,
        bold=True,
        italic=False,
        strike=None,
        underline=None,
    )
   # top row

    font2 = Font(
        name="Calibri",
        size=11,
        bold=False,
        italic=False,
        strike=None,
        underline=None,
    )
    # later row


    side1 = Side(style="thin",color="000000")
    # border width thin , color black.
    columns = ('A','B','C','D','E','F','G','H','I')
    # that's just column use A to I.

    for col in columns:

        ws.column_dimensions[f"{col}"].width = 20
        ws.row_dimensions[1].height = 25
        ws.row_dimensions[2].height = 67

        for i in range(1,ws.max_row+1):
            # get max row_num, row 1 to max change style.
            if 'D' in col:
                if i > 2:
                    cpus = ws["D" + str(i)].value
                    cpu = percent_to_int(cpus)


                    if cpu > 85:
                        ws[f'{col}{i}'].fill = PatternFill(patternType="solid", fgColor="FF3300")
                    elif cpu > 60:
                        ws[f'{col}{i}'].fill = PatternFill(patternType="solid", fgColor="DA9694")
                    elif cpu > 35:
                        ws[f'{col}{i}'].fill = PatternFill(patternType="solid", fgColor="FABF8F")


            if 'E' in col:
                if i > 2:
                    memorys = ws["E" + str(i)].value
                    memory = percent_to_int(memorys)


                    if memory > 85:
                        ws[f'{col}{i}'].fill = PatternFill(patternType="solid", fgColor="FF3300")
                    elif memory > 60:
                        ws[f'{col}{i}'].fill = PatternFill(patternType="solid", fgColor="DA9694")
                    elif memory > 35:
                        ws[f'{col}{i}'].fill = PatternFill(patternType="solid", fgColor="FABF8F")

            if 'F' in col:
                if i > 2:
                    uptimes = ws["F" + str(i)].value
                    uptime = percent_to_int(uptimes)

                    if uptime < 1:
                        ws[f'{col}{i}'].fill = PatternFill(patternType="solid", fgColor="FF3300")



            if i == 1:
                # the top 1 row background color different with other.
                ws[f'{col}{i}'].font = font1
                ws[f'{col}{i}'].fill = PatternFill(patternType="solid",fgColor="FFFF00")
                ws[f'{col}{i}'].alignment = Alignment(horizontal="center", vertical="center" , wrap_text=True)
                cells = ws[f'{col}{i}']
                cells.border = Border(left=side1,right=side1,top=side1,bottom=side1)



            ws[f'{col}{i}'].font = font2
            ws[f'{col}{i}'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cells = ws[f'{col}{i}']
            cells.border = Border(left=side1,right=side1,top=side1,bottom=side1)

    ws.freeze_panes = 'I3'
    wb.save(file_name)
    # end close the Wrokbook.



def create_excel(sheet_name,file_name):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(['IP', 'Model', 'Name', 'CPU', 'Memory', 'Uptime', 'Version', 'Package', 'Device'])
    ws.append(['\nDevice Manage IP\n',
               '\nDevice model\n',
               '\nDevice Sysname\n',
               '\nCPU usage\n',
               '\nMemory usage\n',
               '\nDevice online time\nin weeks\n',
               '\nDevice software \nVersion\n',
               '\nDevice package information\n',
               '\nDevice module \nstatus\n'])
    wb.save(file_name)


def Check(ips,type,username,password):

    ip1 = ips.rstrip()
    ip = str(ip1)


    #date = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M")
    date = datetime.datetime.now().strftime("%Y-%m-%d")

    path = 'output/'+date

    try:
        ssh_client = paramiko.SSHClient()
        ssh_client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        ssh_client.connect(hostname=ip, username=username, password=password,timeout=60)
        #初始化 paramiko 并登录设备




        command = ssh_client.invoke_shell()
        print('Success Connetion ', ip)


        if str("Switch") in str(type):
            out = ""
            cmds = ('display version',
                    'display cpu-u',
                    'display memory-usage',
                    'display patch-information',
                    'display device',
                    'display current-configuration')

            for cmd in cmds:
                command.send(cmd+"\n")
                time.sleep(0.5)

                while True:
                    page = command.recv(65535)
                    page = page.decode("ASCII")
                    out += page
                    time.sleep(0.1)
                    if page.endswith('>') or page.endswith(']'):
                        # if page[-1] in ['>',']']:
                        break
                    if "  ---- More ----" in page:
                        command.send(" ")



            out += command.recv(6553500)

            with open(path+ '/' + ip + '_' + date + '_temp.txt', 'w') as f:
                f.write(out)
                f.close()

        elif str("Firewall") in str(type):
            out = ""
            cmds = ('display version',
                    'display cpu-u',
                    'display memory-usage',
                    'display patch-information',
                    'display device',
                    'display current-configuration')

            for cmd in cmds:
                command.send(cmd+"\n")
                time.sleep(0.5)
                while True:
                    page = command.recv(65535)
                    page = page.decode("ASCII")
                    out += page
                    time.sleep(0.1)
                    if page.endswith('>') or page.endswith(']'):
                        # if page[-1] in ['>',']']:
                        break
                    if "  ---- More ----" in page:
                        command.send(" ")

            out += command.recv(819200)

            with open(path + '/' + ip + '_' + date + '_temp.txt', 'w') as f:
                f.write(out)
                f.close()

        else:
            print()



        ssh_client.close()

        with open(path + '/' + ip + '_' + date + '_temp.txt','r') as fr,open(path + '/' + ip + '_' + date + '.txt', 'w') as fw:
            for text in fr.readlines():
                if text.split():
                    fw.write(text)

    except Exception as f:
        logging.error(ip+f)


def getip():
    ip_list    = []
    type_list  = []
    user_list  = []
    paswd_list = []





    wb1 = load_workbook('Networklist.xlsx')

    ws1 = wb1['Sheet1']

    for i in range(3,ws1.max_row+1):

        IP   = ws1["b"+str(i)].value
        TYPE = ws1["c"+str(i)].value
        USER = ws1["h"+str(i)].value
        PAWD = ws1["i"+str(i)].value


        ip_list.append(IP)
        type_list.append(TYPE)
        user_list.append(USER)
        paswd_list.append(PAWD)

    return ip_list,type_list,user_list,paswd_list
    # return must informaiton

def send_message(msg):
    msg = str(msg)
    nums = ('17786504924','15623006417')
    for num in nums:
        i = requests.get(
            f'http://172.16.54.194:30800/esb-scheduler/api/Message_Info?account=whtk&passwd=tkyy123&phone={num}&content={msg}&uid=F84FFE14-0ACE-4649-BB7D-D09C83923A6A')
        if '200' in str(i):
            print('msg sent!')
        else:
            print('msg faild!')


def verify_IP():
    date = datetime.datetime.now().strftime("%Y-%m-%d")
    path = 'output/'+date


    os.system(f'rm output/{date}/*temp* -rf')

    z = 1
    wb1 = openpyxl.load_workbook('Networklist.xlsx')
    ws1 = wb1['Sheet1']
    f = open(path + '/' + 'fault.txt', 'a')
    for i in range(3, ws1.max_row + 1):


        hosts = ws1['b'+str(i)].value
        #hosts = host.rstrip()

        try:
            open(path + '/' + str(hosts) + '_' + date + '.txt', 'r')
            f.write(str(hosts)+' Connetion!\n')


        except IOError as err:
            print (str(hosts)+' No_Connetion!')
            f.write(str(hosts)+' No_Connetion!\n')
            z +=1

    if z ==1:
        print(f'Backup Complate ready send email! {date}')
        send_message(f'{date} Switch Backups have completed')
        return z

    else:
        send_message(f'{date} Switch backups failed')

    f.close()

def email(date,zip_name):

    fromaddr = '17786504924@189.cn'
    toaddr   = ['example@gmail.com','example@189.cn']




    msg = MIMEMultipart()
    msg['From'] = formataddr(['Automatic_Rbot',fromaddr])
    msg['To'] = ','.join(toaddr)
    msg['Subject'] = 'Network Automation Backups Daily ：'+date



    body = ' \nAll basic network equipment, configuration backup has been completed.\n The attachment is the configuration backup, please refer to it! '




    Bckup = zip_name
    Bckups = MIMEApplication(open(Bckup, 'rb').read())
    Bckups.add_header('Content-Disposition', 'attachment', filename='WHTKYY.NetBackAttachment.zip')

    xlsx  = 'output/switch_information_' + date + '.xlsx'
    xlsxs = MIMEApplication(open(xlsx, 'rb').read())
    xlsxs.add_header('Content-Disposition', 'attachment', filename='switch_information.xlsx')

    msg.attach(MIMEText(body, 'plain'))
    msg.attach(Bckups)
    msg.attach(xlsxs)
    try:
        server = smtplib.SMTP('smtp.189.cn', '587')
        server.starttls()
        server.login('youUsername', 'youPassword')
        print('\n Start send the email , loding ....')
        server.sendmail(fromaddr, toaddr, msg.as_string())
        server.quit()
        print('\n E-mail send Successfully ...')
    except smtplib.SMTPException as e:
        print('errors', e)



def get_zip_file(input_path, result):
    files = os.listdir(input_path)
    for file in files:
        if os.path.isdir(input_path + '/' + file):
            get_zip_file(input_path + '/' + file, result)
        else:
            result.append(input_path + '/' + file)

def zip_file_path(input_path, output_path, output_name):
    verify_path = os.path.exists(output_path)
    if not verify_path:
        os.makedirs(output_path)
    f = zipfile.ZipFile(output_path + '/' + output_name, 'w', zipfile.ZIP_DEFLATED)
    filelists = []
    get_zip_file(input_path, filelists)
    for file in filelists:
        f.write(file)
    f.close()
    return output_path + r"/" + output_name




def fillter_data():

    date =datetime.date.today().isoformat()

    wb1 = openpyxl.load_workbook('Networklist.xlsx')
    ws1 = wb1['Sheet1']
    for i in range(3, ws1.max_row + 1):
        ip = ws1['b' + str(i)].value

        try:
            output = open('output/'+date+'/'+ str(ip) + '_' + date + '.txt', 'r')
            output = output.read()
            Models = re.findall(r'\s(\S+\d+).+uptime', str(output))
            if str('[]') in str(Models):
                Model = ['NA', 'NULL']
            else:
                Model = Models

            CPUs = re.findall(r'Usage\D+\s(\S+\%)\s', str(output))
            if str('[]') in str(CPUs):
                CPU = ['NA', 'NULL']
            else:
                CPU = CPUs

            Memeorys = re.findall(r'Mem.+P.+\s(\d+\S)', str(output))
            if str('[]') in str(Memeorys):
                Memeory = ['NA', 'NULL']
            else:
                Memeory = Memeorys

            Uptimes = re.findall(r'time\sis\s(.+day)', str(output))
            if str('[]') in str(Uptimes):
                Uptime = ['NA', 'NULL']
            else:
                Uptime = Uptimes

            #(V\d{1,3}R\d{1,3}C\d{1,2}SPC\d{1,3})
            Versions = re.findall(r'(V\d{1,3}R\d{1,3}C\S+)\S+', str(output))
            if str('[]') in str(Versions):
                Version = ['NA', 'NULL']
            else:
                Version = Versions

            Packages = re.findall(r'(V\d{1,3}R\d{1,3}SPH\d{1,3})', str(output))
            if str('[]') in str(Packages):
                Package = ['NA', 'NULL']
            else:
                Package = Packages

            Sysnames = re.findall(r'ysname\s(.+)', str(output))
            if str('[]') in str(Sysnames):
                Syename = ['NA', 'NULL']
            else:
                Syename = Sysnames

            Devices = re.findall(r'Abnormal', str(output))
            if str('Abnormal') in str(Devices):
                Device = 'Abnormal'
            else:
                Device = 'Normal'


            file_name = 'output/' + 'switch_information_' + date + '.xlsx'
            sheet_name = 'Sheet1'

            wb = openpyxl.load_workbook(file_name)
            ws = wb.active
            ws.title = sheet_name
            try:
                ws.append([ip, Model[0], Syename[0], CPU[0], Memeory[0], Uptime[0], Version[0], Package[0], Device])
            except:
                ws.append([ip])
            wb.save(file_name)
        except:
            pass
