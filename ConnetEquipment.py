#-*-coding:gb2312-*-
from multiprocessing import Pool
from concurrent.futures import ThreadPoolExecutor
import openpyxl
import os
import paramiko
import time
import datetime
import socket
from openpyxl.styles import Font,PatternFill,Alignment,Side,Border


Today = datetime.date.today().isoformat()

def CreateExcel():
    SheetName = 'Sheet0'
    FileName  = 'DeviceLoginInformaiton.xlsx'
    wb =openpyxl.Workbook()
    ws = wb.active
    ws.title = SheetName
    ws.append(['IP', 'Hostname', 'Username', 'Password', 'CommandFile','Model'])
    ws.append(['\nDevice Manage IP\n',
               '\nDevice Hostname\n',
               '\nDevice Login user\n',
               '\nDevice Login password\n',
               '\nInput command file name\n',
               '\nEquipment detailed model \n'])
    ws.append(['172.16.1.1','R1','admin','Admin@123','R1_Config_command.txt','S5700S'])
    ws.append(['172.16.1.2','R2','admin','Admin@123','R2_Config_command.txt','2910'])
    wb.save(FileName)
    ExcelStyle(SheetName,FileName)


    if not os.path.exists(f'output/{Today}'):
        os.makedirs(f'output/{Today}')
    if not os.path.exists(f'result/{Today}'):
        os.makedirs(f'result/{Today}')
    if not os.path.exists('commands'):
        os.makedirs('commands')

    with open('commands/R1_Config_command.txt','w+') as Example:
        Example.write('show arp,3')
        Example.close()

    with open('commands/R2_Config_command.txt','w+') as Example:
        Example.write('show arp,3')
        Example.close()

def ExcelStyle(SheetName,FileName):
    wb = openpyxl.load_workbook(FileName)
    ws = wb.active
    ws.title = SheetName

    font1 = Font(name="Calibri",size=11,bold=True,italic=False,strike=None,underline=None,)
    # top row
    font2 = Font(name="Calibri",size=11,bold=False,italic=False,strike=None,underline=None,)
    # later row
    side1 = Side(style="thin",color="000000")
    # border width thin , color black.
    columns = ('A','B','C','D','E','F')
    # that's just column use A to F.

    for col in columns:
        ws.column_dimensions[f"{col}"].width = 28
        ws.row_dimensions[1].height = 25
        ws.row_dimensions[2].height = 67

        for i in range(1,ws.max_row+1):
            # get max row_num, row 1 to max change style.
            if i == 1:
                # the top 1 row background color different with other.
                ws[f'{col}{i}'].font = font1
                ws[f'{col}{i}'].fill = PatternFill(patternType="solid",fgColor="FFFF00")
                ws[f'{col}{i}'].alignment = Alignment(horizontal="center", vertical="center" , wrap_text=True)
                cells = ws[f'{col}{i}']
                cells.border = Border(left=side1,right=side1,top=side1,bottom=side1)

            ws[f'{col}{i}'].font = font2
            ws[f'{col}{i}'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cells = ws[f'{col}{i}']
            cells.border = Border(left=side1,right=side1,top=side1,bottom=side1)
    ws.freeze_panes = 'F3'
    wb.save(FileName)

def GetConnectInf():

    if not os.path.exists(f'output/{Today}'):
        os.makedirs(f'output/{Today}')
    if not os.path.exists(f'output/temp'):
        os.makedirs(f'output/temp')
    if not os.path.exists(f'result/{Today}'):
        os.makedirs(f'result/{Today}')
    if not os.path.exists('commands'):
        os.makedirs('commands')

    ConnetInf_list = []

    wb = openpyxl.load_workbook('DeviceLoginInformaiton.xlsx')
    ws = wb['Sheet0']
    for i in range(3,ws.max_row+1):
        ip_address  = ws["a"+str(i)].value
        hostname    = ws["b"+str(i)].value
        username    = ws["c"+str(i)].value
        password    = ws["d"+str(i)].value
        commandFile = ws["e"+str(i)].value

        Inf  = ip_address,hostname,username,password,commandFile

        ConnetInf_list.append(Inf)

    return ConnetInf_list

def ConnetionDevice(Inf_List):
    time.sleep(1)
    ip,hostname,username,password,commandFile = Inf_List[0],Inf_List[1],Inf_List[2],Inf_List[3],Inf_List[4]

    try:
        ssh_client = paramiko.SSHClient()
        ssh_client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        ssh_client.connect(hostname=ip, username=username, password=password,timeout=10)


        conn = ssh_client.invoke_shell()
        print(f'Successfuly to  {ip}')
        commands = open('commands/'+commandFile,'r')
        for cmds in commands.readlines():
            cmd = cmds.split(',')
            command,timeout = cmd[0],cmd[1]
            conn.send(command+'\n')
            time.sleep(int(timeout))

        out = b''
        out += conn.recv(819200)



        FileName1 = f'output/temp/' + ip+'_'+hostname+ '_' + Today + '_temp.txt'
        FileName2 = f'output/{Today}/' + ip +'_'+hostname+ '_' + Today + '.txt'

        with open(FileName1, 'wb') as f:
            f.write(out)
            f.close()

        with open(FileName1,'r') as fr,open(FileName2, 'w') as fw:
            for text in fr.readlines():
                if text.split():
                    fw.write(text)


    except paramiko.AuthenticationException:
        AuthError = open(f'result/{Today}/auth_error.txt','a')
        AuthError.write(ip+'_AuthFaild\n')
        AuthError.close()

    except socket.error:
        Unreach = open(f'result/{Today}/unreach_error.txt', 'a')
        Unreach.write(ip+'_Unreach\n')
        Unreach.close()

    except Exception as e:
        Error = open(f'result/{Today}/error.txt', 'a')
        Error.write(ip + '_error\n')
        Error.close()


Banner = '''
-------------------------------------------+
--------Xsa123MKSAMDK12313213123-----------+
--根据需求输入序号---------------------------+
1，首次使用，生成目录以及存放设备信息的Excel    +
2，单线程执行程序 配置命令建议单线程            +
3，多线成执行程序                            +
+------------------------------------------+
'''





if __name__ == '__main__':
    print(Banner)
    Num = int(input('Please you chose: '))

    if Num == 1:
        CreateExcel()

        input('Please enter quit.')

    elif Num ==2:
        Inf = GetConnectInf()
        with ThreadPoolExecutor(max_workers=1) as P:
            P.map(ConnetionDevice,Inf)

        input('Please enter quit.')

    elif Num ==3:
        Inf = GetConnectInf()
        with ThreadPoolExecutor(max_workers=100) as P:
            P.map(ConnetionDevice,Inf)

        input('Please enter quit.')

    else:
        print ('Input error quit.')
        input('Please enter quit.')
